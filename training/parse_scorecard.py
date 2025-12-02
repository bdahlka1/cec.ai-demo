import openpyxl
from pathlib import Path

# Mapping of scorecard row numbers to human-friendly names
CRITERIA = {
    6:  "Approved integrator",
    7:  "Bidder list clarity",
    8:  "Integrator competition",
    9:  "Client/GC relationship",
    11: "SCADA definition",
    12: "PLC brand",
    13: "SCADA brand",
    14: "Instrumentation definition",
    15: "Schedule realism",
    22: "Geography",
    23: "Bid timing",
    24: "Strategic value",
    25: "Liquidated damages",
    26: "Delivery model",
    27: "Installation responsibility",
}

ROWS = list(CRITERIA.keys())

def load_scorecard(scorecard_path: str):
    """
    Loads a historical XLSX scorecard and extracts:
      - total score (B35)
      - points & comments for each criterion row
    """
    path = Path(scorecard_path)
    wb = openpyxl.load_workbook(path, data_only=True)
    ws = wb.active

    criteria_data = {}
    for row in ROWS:
        points = ws.cell(row=row, column=6).value or 0   # Column F
        comments = ws.cell(row=row, column=7).value or ""  # Column G
        name = CRITERIA[row]

        criteria_data[name] = {
            "row": row,
            "points": float(points),
            "comment": str(comments).strip()
        }

    total_score = float(ws["B35"].value or 0)

    return {
        "total_score": total_score,
        "criteria": criteria_data,
    }
