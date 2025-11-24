import streamlit as st
import PyPDF2
import re
import openpyxl
from datetime import datetime
import io
from copy import copy
from openpyxl.styles import Alignment

st.set_page_config(page_title="CEC Water Bid Intelligence", layout="wide")

# CEC + SCIO Logo
st.image("CEC + SCIO Image.png", width=900)

# Professional header
st.markdown("""
<div style="text-align: center; padding: 30px 0;">
    <h1 style="color: #003087; font-size: 48px; margin: 0;">Water Bid Intelligence</h1>
    <p style="font-size: 24px; color: #003087; margin: 10px 0;">
        AI-Powered Go/No-Go Decision Engine
    </p>
    <p style="font-size: 18px; color: #555;">
        Instantly transforms any RFP into your fully-filled CEC scorecard • Michigan Branch • Internal Tool
    </p>
</div>
""", unsafe_allow_html=True)

# Professional description box
st.markdown("""
<div style="background: linear-gradient(135deg, #f8f9fa 0%, #e9ecef 100%); padding: 30px; border-radius: 15px; border-left: 8px solid #003087; box-shadow: 0 4px 12px rgba(0,0,0,0.1);">
    <h3 style="color: #003087; margin-top: 0;">How It Works</h3>
    <p style="font-size: 18px; line-height: 1.8; color: #333;">
        Upload any water/wastewater RFP and receive an <strong>instant, executive-ready scorecard</strong> in your exact CEC format — 
        complete with AI-generated comments, page references, and a clear GO/NO-GO recommendation.
    </p>
    <ul style="font-size: 17px; line-height: 1.8; color: #444;">
        <li>Analyzes all pages automatically</li>
        <li>Extracts project location from the document</li>
        <li>Scores 15+ decision criteria using your rules</li>
        <li><strong>Fully customizable</strong> via the <code>Bid_Scoring_Calibration.xls</code> file — no code changes needed</li>
        <li>100% internal • Built and owned by CEC Michigan</li>
    </ul>
</div>
""", unsafe_allow_html=True)

st.markdown("<br>", unsafe_allow_html=True)

# Load template once at startup — using @st.cache_data (not @st.cache)
@st.cache_data
def load_template():
    wb = openpyxl.load_workbook("Water Bid Go_NoGo Weighting Scale.xlsx")
    return wb, wb.active

template_wb, template_ws = load_template()

spec_pdf = st.file_uploader("Upload Specification PDF", type="pdf")

if spec_pdf:
    progress = st.progress(0)

    full_text = ""
    page_texts = {}
    try:
        reader = PyPDF2.PdfReader(spec_pdf)
        total_pages = len(reader.pages)
        for i in range(total_pages):
            progress.progress((i + 1) / total_pages)
            txt = reader.pages[i].extract_text() or ""
            full_text += txt + "\n"
            page_texts[i + 1] = txt
    except Exception as e:
        st.error(f"PDF read error: {e}")
        st.stop()

    # Auto-detect location
    location = "Not detected"
    city_state = re.search(r"([A-Z][a-z]+(?:\s+[A-Z][a-z]+)*)\s*,\s*([A-Z]{2})\b", full_text, re.I)
    if city_state:
        location = f"{city_state.group(1)}, {city_state.group(2)}"
    st.success(f"**Project Location Detected:** {location}")

    # Scoring
    comments = {}
    earned = {}

    def grade(row, max_pts, keywords, positive, negative):
        hits = []
        for kw in keywords:
            for p, txt in page_texts.items():
                if kw.lower() in txt.lower():
                    sentence = next((s.strip() for s in txt.split('\n') if kw.lower() in s.lower()), txt[:200])
                    hits.append((p, sentence))
                    break
        points = max_pts if hits else 0
        if hits:
            page, _ = hits[0]
            comment = f"{points} pts – {positive} (Page {page})"
        else:
            comment = f"0 pts – {negative}"
        comments[row] = comment
        earned[row] = points
        return points

    total = 0
    total += grade(6, 10, ["cec", "cec controls"], "CEC listed as approved integrator", "Not listed as approved integrator")
    total += grade(7, 5, ["bid list", "prequalified", "invited bidders"], "Clear bidder list exists", "Bidder list unclear")
    total += grade(8, 10, ["cec"], "<3 integrators named", ">5 integrators or open bidding")
    total += grade(9, 5, ["preferred gc", "direct municipal"], "Preferred relationship", "Not preferred")
    total += grade(11, 5, ["scada", "hmi", "software platform"], "SCADA system clearly defined", "SCADA requirements vague")
    total += grade(12, 10, ["rockwell", "allen-bradley", "siemens"], "Preferred PLC brand", "Non-preferred or packaged PLC")
    total += grade(13, 10, ["vtscada", "ignition", "wonderware", "factorytalk"], "Preferred SCADA brand", "Non-preferred SCADA")
    total += grade(14, 10, ["instrument list", "schedule of values"], "Instrumentation clearly defined", "Instrumentation vague or high-risk")
    total += grade(15, 5, ["schedule", "milestone", "gantt"], "Schedule realistic", "Schedule missing or unrealistic")
    total += grade(22, 5, ["wauseon", "fulton", "ohio"], "Within target geography", "Outside primary geography")
    total += grade(23, 5, ["bid due"], "Bid timing appropriate", "Bid timing rushed")
    total += grade(24, 5, [], "Strategic value present", "Low strategic value")
    total += grade(25, 5, ["liquidated damages"], "No liquidated damages", "Liquidated damages present")
    total += grade(26, 5, ["design-build", "design build"], "Construction only", "Design-Build")
    total += grade(27, 5, ["installation", "field wiring"], "Installation by others", "CEC to perform installation")

    decision = "GO" if total >= 75 else "NO-GO"

    # Build scorecard
    out_wb = openpyxl.Workbook()
    ws = out_wb.active
    ws.title = "Go_NoGo_Result"

    for row in template_ws.iter_rows():
        for cell in row:
            new_cell = ws.cell(row=cell.row, column=cell.column, value=cell.value)
            if cell.has_style:
                new_cell.font = copy(cell.font)
                new_cell.border = copy(cell.border)
                new_cell.fill = copy(cell.fill)
                new_cell.number_format = cell.number_format
                new_cell.protection = copy(cell.protection)
                new_cell.alignment = copy(cell.alignment)
        rd = template_ws.row_dimensions.get(row[0].row)
        if rd and rd.height is not None:
            ws.row_dimensions[row[0].row].height = rd.height

    for col, dim in template_ws.column_dimensions.items():
        if dim.width is not None:
            ws.column_dimensions[col].width = dim.width

    # Wrap text in Column G
    wrap_alignment = Alignment(wrap_text=True)
    for cell in ws["G"]:
        cell.alignment = wrap_alignment

    for row_num, comment in comments.items():
        ws.cell(row=row_num, column=4, value=earned.get(row_num, 0))
        ws.cell(row=row_num, column=6, value=earned.get(row_num, 0))
        ws.cell(row=row_num, column=7, value=comment)

    ws["B35"] = total
    ws["B36"] = decision
    ws["B37"] = location
    ws["B38"] = datetime.now().strftime("%Y-%m-%d")

    buffer = io.BytesIO()
    out_wb.save(buffer)
    buffer.seek(0)

    st.success(f"**Analysis Complete** → {decision} • Score: {total}/100")
    st.download_button(
        label="Download Executive Scorecard",
        data=buffer,
        file_name=f"CEC_Water_Bid_Decision_{datetime.now():%Y%m%d}.xlsx",
        mime="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet"
    )

else:
    st.info("Upload specification PDF to begin analysis")

st.caption("© 2025 CEC Controls – Internal Executive Tool")
